"""Document-level quality profiling."""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any

from data_governance.core.config import GovernanceConfig
from data_governance.core.models import (
    DataAsset,
    DataAssetType,
    QualityDimension,
    QualityReport,
    QualityScore,
)
from data_governance.profiler.metrics import QualityMetrics


class DocumentProfiler:
    """Profile quality of document files in the knowledge base."""

    SUPPORTED_EXTENSIONS = {".txt", ".md", ".pdf", ".docx", ".xlsx"}

    def __init__(self, config: GovernanceConfig | None = None):
        self.config = config or GovernanceConfig()
        self.metrics = QualityMetrics()

    def profile_file(self, file_path: str | Path) -> QualityReport:
        """Profile a single document file."""
        path = Path(file_path)
        content = self._read_file(path)

        asset = DataAsset(
            id=f"doc:{path.name}",
            name=path.name,
            asset_type=DataAssetType.DOCUMENT,
            source_path=str(path),
        )

        dimension_scores: list[QualityScore] = []

        # Completeness
        completeness = self.metrics.completeness_score(
            content, min_length=self.config.profiler.min_chunk_length
        )
        dimension_scores.append(
            QualityScore(
                dimension=QualityDimension.COMPLETENESS,
                score=completeness,
                details=f"Content length: {len(content)} chars",
                metrics={"length": len(content), "word_count": self.metrics.word_count(content)},
            )
        )

        # Validity
        gibberish = self.metrics.gibberish_score(content)
        validity = 1.0 - gibberish
        dimension_scores.append(
            QualityScore(
                dimension=QualityDimension.VALIDITY,
                score=validity,
                details="Low corruption" if validity > 0.8 else "Possible corruption detected",
                metrics={
                    "gibberish_score": gibberish,
                    "char_density": self.metrics.char_density(content),
                    "cjk_ratio": self.metrics.cjk_ratio(content),
                },
            )
        )

        # Timeliness
        try:
            mtime = os.path.getmtime(path)
            freshness = self.metrics.freshness_score(
                mtime, datetime.now().timestamp(), self.config.freshness.stale_threshold_days
            )
        except OSError:
            freshness = 0.5
        dimension_scores.append(
            QualityScore(
                dimension=QualityDimension.TIMELINESS,
                score=freshness,
                details=f"Last modified: {datetime.fromtimestamp(mtime).isoformat()}"
                if freshness < 1.0
                else "Recently updated",
                metrics={"file_mtime": mtime if "mtime" in dir() else None},
            )
        )

        overall = sum(s.score for s in dimension_scores) / len(dimension_scores)
        recommendations = self._generate_recommendations(dimension_scores, content)

        return QualityReport(
            asset_id=asset.id,
            asset_name=asset.name,
            asset_type=DataAssetType.DOCUMENT,
            overall_score=overall,
            dimension_scores=dimension_scores,
            total_items=1,
            issues_found=sum(1 for s in dimension_scores if s.score < 0.8),
            recommendations=recommendations,
        )

    def profile_directory(self, dir_path: str | Path) -> list[QualityReport]:
        """Profile all documents in a directory recursively."""
        dir_path = Path(dir_path)
        reports = []
        for path in sorted(dir_path.rglob("*")):
            if path.is_file() and path.suffix.lower() in self.SUPPORTED_EXTENSIONS:
                try:
                    report = self.profile_file(path)
                    reports.append(report)
                except Exception as e:
                    reports.append(
                        QualityReport(
                            asset_id=f"doc:{path.name}",
                            asset_name=path.name,
                            asset_type=DataAssetType.DOCUMENT,
                            overall_score=0.0,
                            issues_found=1,
                            recommendations=[f"Failed to profile: {e}"],
                        )
                    )
        return reports

    def _read_file(self, path: Path) -> str:
        """Read file content as text."""
        suffix = path.suffix.lower()
        if suffix in (".txt", ".md"):
            return path.read_text(encoding="utf-8", errors="replace")
        elif suffix == ".pdf":
            return self._read_pdf(path)
        elif suffix == ".docx":
            return self._read_docx(path)
        elif suffix == ".xlsx":
            return self._read_xlsx(path)
        return ""

    def _read_pdf(self, path: Path) -> str:
        try:
            from pypdf import PdfReader

            reader = PdfReader(str(path))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        except ImportError:
            return path.read_text(encoding="utf-8", errors="replace")

    def _read_docx(self, path: Path) -> str:
        try:
            from docx import Document

            doc = Document(str(path))
            return "\n".join(p.text for p in doc.paragraphs)
        except ImportError:
            return ""

    def _read_xlsx(self, path: Path) -> str:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(path), read_only=True)
            texts = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    texts.append(" ".join(str(c) for c in row if c is not None))
            return "\n".join(texts)
        except ImportError:
            return ""

    def _generate_recommendations(
        self, scores: list[QualityScore], content: str
    ) -> list[str]:
        """Generate actionable recommendations based on scores."""
        recs: list[str] = []
        for s in scores:
            if s.dimension == QualityDimension.COMPLETENESS and s.score < 0.5:
                recs.append("Document is too short or empty — consider enriching content")
            if s.dimension == QualityDimension.VALIDITY and s.score < 0.7:
                recs.append("Possible encoding corruption detected — re-export from source")
            if s.dimension == QualityDimension.TIMELINESS and s.score < 0.3:
                recs.append("Document is stale — review and update or archive")
        if not content.strip():
            recs.append("Document is empty — remove or replace")
        return recs
